"""
EXCEL LOADER - Carga y validación de archivos Excel con contactos
Goleador SMS Marketing - Carga de Campañas Dinámicas
"""

import logging
import os
import uuid
from datetime import datetime
from typing import Dict, List, Tuple
import json

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

from utils import PhoneValidator

logger = logging.getLogger(__name__)

class ExcelLoader:
    """
    Cargador de archivos Excel con validación de contactos y variables dinámicas
    """

    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}
    REQUIRED_COLUMNS = {'numero', 'phone', 'cel', 'número'}

    def __init__(self):
        """Inicializar cargador de Excel"""
        self.phone_validator = PhoneValidator()
        logger.info("✅ ExcelLoader inicializado")

    def read_excel(self, file_path: str, sheet_name: str = 'Contactos') -> Dict:
        """
        Leer archivo Excel y extraer contactos con variables

        Args:
            file_path: Ruta del archivo Excel
            sheet_name: Nombre de la hoja (por defecto 'Contactos')

        Returns:
            Dict con contactos y metadatos
        """
        logger.info(f"📖 Leyendo archivo: {file_path}")

        # Validar archivo
        validation = self.validate_file(file_path)
        if not validation['valid']:
            return {
                "status": "error",
                "message": validation['error'],
                "errors": [validation['error']]
            }

        try:
            # Detectar extensión
            ext = os.path.splitext(file_path)[1].lower()

            if ext == '.xlsx':
                return self._read_xlsx(file_path, sheet_name)
            elif ext == '.xls':
                return self._read_xls(file_path, sheet_name)
            elif ext == '.csv':
                return self._read_csv(file_path)
            else:
                return {
                    "status": "error",
                    "message": f"Formato no soportado: {ext}",
                    "errors": [f"Formato no soportado: {ext}"]
                }

        except Exception as e:
            logger.error(f"❌ Error leyendo archivo: {str(e)}")
            return {
                "status": "error",
                "message": f"Error al leer archivo: {str(e)}",
                "errors": [str(e)]
            }

    def _read_xlsx(self, file_path: str, sheet_name: str) -> Dict:
        """Leer archivo XLSX"""
        if not OPENPYXL_AVAILABLE:
            return {
                "status": "error",
                "message": "openpyxl no está instalado",
                "errors": ["Instalar: pip install openpyxl"]
            }

        try:
            wb = load_workbook(file_path, data_only=True)

            # Intentar encontrar la hoja
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]  # Usar primera hoja
                logger.warning(f"⚠️ Hoja '{sheet_name}' no encontrada, usando: {sheet_name}")

            ws = wb[sheet_name]

            # Extraer encabezados
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip())
                else:
                    headers.append(f"columna_{len(headers) + 1}")

            # Validar columnas requeridas
            phone_col = self._find_phone_column(headers)
            if phone_col is None:
                return {
                    "status": "error",
                    "message": "No se encontró columna de números de teléfono",
                    "errors": ["Debe haber una columna: 'numero', 'phone', 'cel' o 'número'"]
                }

            # Extraer datos
            contacts = []
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                contact = self._parse_row(row, headers, phone_col, row_idx, errors)
                if contact:
                    contacts.append(contact)

            return self._process_results(contacts, errors, len(headers))

        except Exception as e:
            logger.error(f"❌ Error en _read_xlsx: {str(e)}")
            return {
                "status": "error",
                "message": str(e),
                "errors": [str(e)]
            }

    def _read_xls(self, file_path: str, sheet_name: str) -> Dict:
        """Leer archivo XLS (Excel antiguo)"""
        if not XLRD_AVAILABLE:
            return {
                "status": "error",
                "message": "xlrd no está instalado",
                "errors": ["Instalar: pip install xlrd"]
            }

        try:
            book = xlrd.open_workbook(file_path)

            # Intentar encontrar la hoja
            if sheet_name not in book.sheet_names():
                sheet_name = book.sheet_names()[0]

            sheet = book.sheet_by_name(sheet_name)

            # Extraer encabezados
            headers = []
            for cell in sheet.row_values(0):
                headers.append(str(cell).lower().strip() if cell else "")

            # Validar columnas
            phone_col = self._find_phone_column(headers)
            if phone_col is None:
                return {
                    "status": "error",
                    "message": "No se encontró columna de números",
                    "errors": ["Debe haber: 'numero', 'phone', 'cel'"]
                }

            # Extraer datos
            contacts = []
            errors = []

            for row_idx in range(1, sheet.nrows):
                row = sheet.row_values(row_idx)
                contact = self._parse_row(row, headers, phone_col, row_idx + 1, errors)
                if contact:
                    contacts.append(contact)

            return self._process_results(contacts, errors, len(headers))

        except Exception as e:
            logger.error(f"❌ Error en _read_xls: {str(e)}")
            return {
                "status": "error",
                "message": str(e),
                "errors": [str(e)]
            }

    def _read_csv(self, file_path: str) -> Dict:
        """Leer archivo CSV"""
        import csv

        try:
            contacts = []
            errors = []

            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)

                if not reader.fieldnames:
                    return {
                        "status": "error",
                        "message": "CSV vacío",
                        "errors": ["El archivo CSV está vacío"]
                    }

                headers = [h.lower().strip() for h in reader.fieldnames]

                phone_col = self._find_phone_column(headers)
                if phone_col is None:
                    return {
                        "status": "error",
                        "message": "No se encontró columna de números",
                        "errors": ["Debe haber: 'numero', 'phone', 'cel'"]
                    }

                for row_idx, row in enumerate(reader, start=2):
                    contact = self._parse_row(
                        [row.get(h, '') for h in headers],
                        headers,
                        phone_col,
                        row_idx,
                        errors
                    )
                    if contact:
                        contacts.append(contact)

            return self._process_results(contacts, errors, len(headers))

        except Exception as e:
            logger.error(f"❌ Error en _read_csv: {str(e)}")
            return {
                "status": "error",
                "message": str(e),
                "errors": [str(e)]
            }

    def validate_file(self, file_path: str) -> Dict:
        """Validar archivo"""
        # Verificar existencia
        if not os.path.exists(file_path):
            return {"valid": False, "error": "Archivo no encontrado"}

        # Verificar extensión
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in self.ALLOWED_EXTENSIONS:
            return {
                "valid": False,
                "error": f"Formato no permitido: {ext}. Usa: {', '.join(self.ALLOWED_EXTENSIONS)}"
            }

        # Verificar tamaño
        size = os.path.getsize(file_path)
        if size > self.MAX_FILE_SIZE:
            return {
                "valid": False,
                "error": f"Archivo muy grande ({size / 1024 / 1024:.1f}MB). Máximo: 10MB"
            }

        return {"valid": True}

    def _find_phone_column(self, headers: List[str]) -> int:
        """Encontrar índice de columna de teléfono"""
        for idx, header in enumerate(headers):
            if any(keyword in header for keyword in self.REQUIRED_COLUMNS):
                return idx
        return None

    def _parse_row(self, row: tuple, headers: List[str], phone_col: int, row_idx: int, errors: List[str]) -> Dict:
        """Parsear una fila de datos"""
        try:
            # Extraer número
            phone = str(row[phone_col]).strip() if phone_col < len(row) else ""

            if not phone:
                errors.append(f"Fila {row_idx}: número vacío")
                return None

            # Validar y normalizar número
            if not self.phone_validator.validate_number(phone):
                errors.append(f"Fila {row_idx}: número inválido: {phone}")
                return None

            normalized_phone = self.phone_validator.format_number(phone)

            # Extraer variables dinámicas
            variables = {}
            for idx, header in enumerate(headers):
                if idx != phone_col and header and idx < len(row):
                    value = row[idx]
                    if value:
                        variables[header] = str(value).strip()

            # Extraer nombre si existe
            nombre = variables.get('nombre', variables.get('name', ''))

            return {
                "numero": normalized_phone,
                "nombre": nombre,
                "email": variables.get('email', variables.get('correo', '')),
                "variables": variables,
                "row_number": row_idx
            }

        except Exception as e:
            errors.append(f"Fila {row_idx}: Error procesando: {str(e)}")
            return None

    def _process_results(self, contacts: List[Dict], errors: List[str], total_columns: int) -> Dict:
        """Procesar resultados y detectar duplicados"""
        logger.info(f"📊 Procesados {len(contacts)} contactos válidos, {len(errors)} errores")

        # Detectar duplicados
        seen_numbers = set()
        unique_contacts = []
        duplicates = []

        for contact in contacts:
            if contact['numero'] in seen_numbers:
                duplicates.append(contact['numero'])
            else:
                seen_numbers.add(contact['numero'])
                unique_contacts.append(contact)

        if duplicates:
            errors.append(f"⚠️ {len(duplicates)} números duplicados detectados")

        return {
            "status": "success",
            "excel_import_id": str(uuid.uuid4()),
            "total_rows": len(contacts) + len(errors),
            "valid_rows": len(unique_contacts),
            "invalid_rows": len(errors),
            "duplicate_rows": len(duplicates),
            "errors": errors,
            "contacts": unique_contacts,
            "detected_variables": list(set().union(*(c['variables'].keys() for c in unique_contacts))),
            "timestamp": datetime.now().isoformat()
        }


# Instancia global
excel_loader = ExcelLoader()
